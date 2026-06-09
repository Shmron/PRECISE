#!/usr/bin/env python3
"""
Generate DHS codebook Excel — one sheet per recode database.
For each variable: name, DHS standard label, description, type, % non-null, example values.
Wide databases (>500 cols) are filtered to variables present in >=5% of rows.
"""
import duckdb
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re

OUT_EXCEL = "/home/rutendo/PRECISE/DHS_Codebook.xlsx"
BASE = "/home/rutendo/PRECISE"

# ── Recode catalogue ──────────────────────────────────────────────────────────
RECODES = {
    "AR": ("dhs_ar.duckdb",  "hiv_results",         "HIV Test Results Recode",
           "One row per person with a DHS HIV blood test. Links to IR/MR/PR by cluster+household+line."),
    "BR": ("dhs_br.duckdb",  "births",              "Births Recode",
           "One row per birth ever reported by interviewed women. Use for fertility, child mortality, birth-spacing."),
    "CR": ("dhs_cr.duckdb",  "couples",             "Couples' Recode",
           "One row per cohabiting couple where both partners were interviewed. Contraception, fertility, HIV concordance."),
    "GE": ("dhs_ge.duckdb",  "clusters",            "Geographic Data (GPS Clusters)",
           "One row per DHS sample cluster. Contains GPS coordinates, region, urban/rural. Key link: DHSCLUST = v001."),
    "GR": ("dhs_gr.duckdb",  "antenatal_postnatal", "Pregnancy & Postnatal Care Recode",
           "Per-pregnancy ANC/delivery/postnatal details. Includes complication flags (hemorrhage, eclampsia, etc.)."),
    "HR": ("dhs_hr.duckdb",  "households",          "Household Recode",
           "One row per household. Dwelling, assets, water, sanitation, deaths in past year."),
    "HW": ("dhs_hw.duckdb",  "anthropometry",       "Height & Weight (Anthropometry)",
           "WHO z-scores for children under 5 (HAZ, WAZ, WHZ). Links to KR/PR on cluster+household+line."),
    "IR": ("dhs_ir.duckdb",  "women",               "Individual Recode (Women)",
           "One row per woman 15-49. Main women's survey: contraception, fertility, ANC, child health, attitudes."),
    "KR": ("dhs_kr.duckdb",  "children",            "Children's Recode",
           "One row per child under 5 (last ~5 years). Immunisation, nutrition, illness, anthropometry."),
    "MR": ("dhs_mr.duckdb",  "men",                 "Men's Recode",
           "One row per man (15-49 or 15-59) interviewed. Subset of households only."),
    "NR": ("dhs_nr.duckdb",  "pregnancies",         "Pregnancies Recode",
           "One row per pregnancy (live births, stillbirths, miscarriages). Links to BR via pidxb."),
    "PR": ("dhs_pr.duckdb",  "persons",             "Household Member Recode",
           "One row per household member. Roster + education, anthropometry, anaemia for all members."),
    "SR": ("dhs_sr.duckdb",  "siblings",            "Siblings Recode",
           "One row per sibling of female respondent. Used for adult/maternal mortality estimates."),
    "WI": ("dhs_wi.duckdb",  "wealth",              "Wealth Index",
           "Pre-computed household wealth quintiles from PCA on assets. Links via v001+v002."),
    "BR_GEO": ("dhs_br_geo.duckdb", "br_geo",       "Births + GPS + Outcomes (PRECISE)",
               "Births (BR) linked to GPS coordinates (GE), birth outcomes (IR), pregnancy complications (GR), NR outcomes. Main analysis-ready table."),
}

# ── Standard DHS variable label dictionary ────────────────────────────────────
LABELS = {
    # Identification / sampling
    "survey_code":  "Survey code (country + recode + phase, e.g. KEBR8C)",
    "country_code": "ISO2 country code",
    "country_name": "Country name",
    "dhs_phase":    "DHS phase (1–8)",
    "source_file":  "Source .dta filename",
    "caseid":       "Woman's unique case ID within survey",
    "hhid":         "Household ID",
    "whhid":        "Wealth index household ID",
    "bidx":         "Birth index (1 = most recent birth)",
    "bidx97":       "Birth index (97-code)",
    "bidxp":        "Birth index for previous interval",
    "bord":         "Birth order number",
    "pidx":         "Pregnancy index (1 = most recent pregnancy)",
    "pidxb":        "Birth history index — links NR pregnancy to BR birth record",
    "pord":         "Pregnancy order",
    # v-series: location / interview
    "v001": "Cluster number (= DHSCLUST in GE)",
    "v002": "Household number",
    "v003": "Respondent line number",
    "v004": "Ultimate area unit",
    "v005": "Sample weight (divide by 1,000,000 for probability weight)",
    "v006": "Month of interview",
    "v007": "Year of interview",
    "v008": "Date of interview (century-month code, CMC)",
    "v009": "Respondent month of birth",
    "v010": "Respondent year of birth",
    "v011": "Respondent date of birth (CMC)",
    "v012": "Respondent current age (years)",
    "v013": "Age group (5-year bands)",
    "v014": "Completeness of month of birth",
    "v015": "Interview result (1=completed)",
    "v016": "Day of interview",
    "v017": "Day of birth",
    "v018": "Completeness of day of birth",
    "v019": "Completeness of day of interview",
    "v020": "Household member status",
    "v021": "Primary sampling unit (PSU)",
    "v022": "Sample stratum",
    "v023": "Stratification variable (for SEs)",
    "v024": "Region",
    "v025": "Urban/rural (1=urban, 2=rural)",
    "v026": "Household location",
    "v027": "Interviewer number",
    "v028": "Supervisor number",
    "v030": "Field check total",
    "v031": "Record type",
    "v032": "Questionnaire number",
    # v-series: background
    "v040": "Cluster altitude",
    "v101": "Region (de jure)",
    "v102": "Type of place of residence",
    "v103": "Childhood place of residence",
    "v104": "Years lived in current place",
    "v105": "Previous place of residence type",
    "v106": "Highest education level (0=none, 1=primary, 2=secondary, 3=higher)",
    "v107": "Highest year of education",
    "v108": "Literacy",
    "v109": "Reading ability",
    "v113": "Source of drinking water",
    "v115": "Time to get water",
    "v116": "Type of toilet facility",
    "v119": "Has electricity",
    "v120": "Has radio",
    "v121": "Has television",
    "v122": "Has refrigerator",
    "v123": "Has bicycle",
    "v124": "Has motorcycle",
    "v125": "Has car/truck",
    "v127": "Main floor material",
    "v128": "Main wall material",
    "v129": "Main roof material",
    "v130": "Religion",
    "v131": "Ethnicity / tribe",
    "v133": "Education in single years",
    "v134": "De facto place of residence",
    "v135": "Current residence status",
    "v136": "Number of household members",
    "v137": "Number of children under 5 in household",
    "v138": "Number of eligible women in household",
    "v139": "Residence for weighting",
    "v140": "Altitude",
    "v149": "Educational attainment",
    "v150": "Relationship to household head",
    "v151": "Sex of household head",
    "v152": "Age of household head",
    "v153": "Has mobile phone",
    "v155": "Literacy (self-report)",
    "v157": "Reads newspaper",
    "v158": "Listens to radio",
    "v159": "Watches television",
    "v160": "Has clock",
    "v161": "Has refrigerator",
    "v162": "Has TV",
    "v163": "Has computer",
    "v164": "Has landline phone",
    "v169a": "Owns mobile phone",
    "v190": "Wealth index quintile (1=poorest, 5=richest)",
    "v190a": "Wealth quintile (urban/rural-stratified)",
    "v191": "Wealth index factor score (continuous)",
    # v-series: fertility / reproductive
    "v201": "Total children ever born",
    "v202": "Sons at home",
    "v203": "Daughters at home",
    "v204": "Sons elsewhere",
    "v205": "Daughters elsewhere",
    "v206": "Sons who have died",
    "v207": "Daughters who have died",
    "v208": "Births in last 5 years",
    "v209": "Births in last 3 years",
    "v210": "Births in last year",
    "v211": "Date of first birth (CMC)",
    "v212": "Age at first birth",
    "v213": "Currently pregnant (0=no, 1=yes)",
    "v214": "Duration of current pregnancy (months)",
    "v215": "Months since last terminated pregnancy",
    "v216": "Month of last menstrual period",
    "v217": "Knowledge of fertile period",
    "v218": "Living children at interview",
    "v219": "Living children + current pregnancy",
    "v220": "Ideal number of children",
    "v221": "Elapsed time since last menstrual period",
    "v222": "Months since last birth",
    "v223": "Months since last birth (imputed)",
    "v224": "Number of ANC visits (last birth)",
    "v225": "Wanted last pregnancy (1=wanted then, 2=wanted later, 3=not wanted)",
    "v226": "Duration of breastfeeding last child",
    "v228": "Had a pregnancy that did not end in live birth (1=yes)",
    "v229": "Visited health facility in last 12 months for FP",
    "v230": "Told about FP at health facility",
    "v234": "Obstetric fistula (1=yes)",
    "v237": "Postnatal check (baby) within 2 months",
    "v238": "Number of ANC visits (last birth — summary)",
    "v239": "Postnatal check timing",
    "v240": "Postnatal care provider",
    "v241": "Year of last birth",
    "v242": "Date of last birth (CMC)",
    "v243": "Delivery assistance at last birth",
    "v244": "Tetanus injection during last pregnancy",
    "v245": "Iron supplements during last pregnancy",
    "v246": "Vitamin A supplements after last birth",
    "v248": "Blood pressure taken during ANC",
    "v249": "Months pregnant at last ANC visit",
    # v-series: contraception
    "v301": "Knowledge of any contraceptive method",
    "v302": "Knowledge of any modern method",
    "v312": "Current contraceptive method used",
    "v313": "Current use of contraception (broad category)",
    "v317": "Date of start of current contraceptive use (CMC)",
    "v321": "Use of contraceptive in last 5 years",
    "v323": "Brand of oral pill used",
    "v325": "Source of current method",
    "v337": "Month first method was used",
    "v364": "Contraceptive use and intention",
    "v367": "Wanted last child",
    "v369": "Use of contraception after birth",
    "v370": "Reason not using contraception",
    "v375a": "Main reason for not using contraception",
    "v376": "Reason for method discontinuation (last 5 years)",
    "v381": "Source of method (current user)",
    "v384a": "Exposure to FP messages on radio",
    "v384b": "Exposure to FP messages on TV",
    "v384c": "Exposure to FP messages in newspaper",
    "v393": "Visited by FP worker in last 12 months",
    "v394": "Visited health facility in last 12 months",
    "v395": "Told about FP at facility visit",
    # b-series: birth history
    "b0": "Twin/triplet indicator (0=singleton, 1=first twin, 2=second twin)",
    "b1": "Month of birth",
    "b2": "Year of birth",
    "b3": "Date of birth (CMC — century-month code)",
    "b4": "Sex of child (1=male, 2=female)",
    "b5": "Child alive at time of survey (0=dead, 1=alive)",
    "b6": "Age at death (century-month days)",
    "b7": "Age at death (months; 0 = died in first month / neonatal)",
    "b8": "Current age (years, if alive)",
    "b9": "Child lives with whom (0=with mother, 1=elsewhere, 2=dead)",
    "b10": "Completeness of birth month",
    "b11": "Preceding birth interval (months since prior birth)",
    "b12": "Succeeding birth interval (months until next birth)",
    "b13": "Flag for birth interval",
    "b14": "Entry into marriage relative to birth",
    "b15": "Exposure at time of birth",
    "b16": "Child line number in household",
    "b17": "Type of birth (singleton/twin)",
    "b18": "Date of death (CMC, if died)",
    "b19": "Current age in months (if alive)",
    "b20": "Gestation/size at birth (phase 7+; 9=DK)",
    "b21": "Month of death (CMC)",
    # m-series: per-birth maternal/delivery care (indexed _1 to _6)
    "m1":   "Month of first antenatal visit",
    "m2":   "Number of antenatal visits",
    "m3a":  "Delivery assistance: doctor (0=no, 1=yes)",
    "m3b":  "Delivery assistance: nurse/midwife (0=no, 1=yes)",
    "m3c":  "Delivery assistance: auxiliary midwife",
    "m3d":  "Delivery assistance: traditional birth attendant",
    "m3e":  "Delivery assistance: community health worker",
    "m3f":  "Delivery assistance: relatives/friends",
    "m3g":  "Delivery assistance: no one",
    "m3h":  "Delivery assistance: other",
    "m3n":  "Delivery assistance: other (alt)",
    "m4":   "Postnatal check for mother (0=no, 1=yes)",
    "m5":   "Duration of amenorrhoea after birth (months)",
    "m6":   "Duration of breastfeeding (months)",
    "m7":   "Duration of post-birth abstinence (months)",
    "m8":   "Duration of abstinence (imputed)",
    "m9":   "In birth interval: amenorrhoeic (0=no, 1=yes)",
    "m10":  "In birth interval: breastfeeding (0=no, 1=yes)",
    "m11":  "In birth interval: abstaining (0=no, 1=yes)",
    "m12":  "Method used during birth interval",
    "m13":  "ANC provider type (0=none, 1=doctor, 2=nurse/midwife, 3=aux. nurse, 6=TBA, 7=other)",
    "m14":  "Months pregnant at first ANC visit (98=DK, 99=missing; 0=not 1st trimester)",
    "m15":  "Place of delivery (11=home, 12=other home, 21=gov hospital, 22=gov health centre, 31=private hospital)",
    "m16":  "Place of delivery (grouped)",
    "m17":  "C-section delivery (0=no, 1=yes)",
    "m17a": "C-section scar seen",
    "m18":  "Source of birth weight info (1=card, 2=maternal recall, 8=DK)",
    "m19":  "Birth weight in grams (9990–9999 = missing/DK; <2500 = low birth weight)",
    "m19a": "Birth weight in kilograms (alternative variable, some surveys)",
    "m27":  "ANC content: blood pressure taken",
    "m28":  "ANC content: urine sample taken",
    "m29":  "ANC content: blood sample taken",
    "m34":  "ANC content: given iron tablets/syrup",
    "m35":  "ANC content: given intestinal parasite drugs",
    "m36":  "ANC content: given malaria prevention",
    "m38":  "ANC content: given tetanus toxoid injection",
    "m39":  "ANC content: examined for HIV",
    "m43":  "Postnatal check for baby (0=no, 1=yes)",
    "m44":  "When postnatal check for baby (days)",
    "m45":  "Perceived size at birth (1=very large, 2=larger than avg, 3=average, 4=smaller, 5=very small)",
    "m46":  "Child given Vitamin A",
    "m54":  "Blood sample taken during last ANC visit",
    "m55":  "Complication after delivery (0=no, 1=yes)",
    "m60":  "Received blood transfusion during delivery (0=no, 1=yes)",
    "m61":  "Place of postnatal check for baby",
    "m62":  "Timing of postnatal check for baby",
    "m63":  "Postnatal check provider (baby)",
    "m64":  "Postnatal check provider type (baby)",
    "m66":  "Postpartum haemorrhage (0=no, 1=yes)",
    "m67":  "Place of postnatal check for mother",
    "m68":  "Timing of postnatal check for mother",
    "m69":  "Postnatal check provider (mother)",
    "m70":  "Received treatment for delivery complication (0=no, 1=yes)",
    "m71":  "Place of treatment for complication",
    "m72":  "Type of complication treated",
    "m73":  "Type of delivery complication treated (alt)",
    "m74":  "Received postnatal check (0=no, 1=yes)",
    "m75":  "Place of postnatal check",
    "m76":  "Type of postnatal check",
    "m77":  "Any complication during delivery (0=no, 1=yes)",
    "m77a": "Number of delivery complications",
    "m78a": "Complication: excessive bleeding / hemorrhage (0=no, 1=yes)",
    "m78b": "Complication: convulsions / eclampsia (0=no, 1=yes)",
    "m78c": "Complication: prolonged or obstructed labour (0=no, 1=yes)",
    "m78d": "Complication: fever / infection (0=no, 1=yes)",
    "m78e": "Complication: obstetric fistula symptoms (0=no, 1=yes)",
    "m78f": "Complication: premature rupture of membranes / PROM (0=no, 1=yes)",
    "m78g": "Complication: malaria during pregnancy (0=no, 1=yes)",
    "m78h": "Complication: anaemia (0=no, 1=yes)",
    "m78i": "Complication: other (0=no, 1=yes)",
    "m78j": "Complication: hypertension / high blood pressure (0=no, 1=yes)",
    "m78k": "Complication: cord prolapse (0=no, 1=yes)",
    "m78l": "Complication: malpresentation (0=no, 1=yes)",
    "m78m": "Complication: pre-eclampsia / gestational hypertension (0=no, 1=yes)",
    "m78n": "Complication: severe bleeding before delivery (antepartum haemorrhage, 0=no, 1=yes)",
    "m78o": "Complication: other specific complication (0=no, 1=yes)",
    "m78p": "Complication: additional complication (country-specific)",
    "m80":  "Postnatal care timing (1=same day, 2=1-2 days, 3=3-6 days, 4=7-41 days, 5=42+ days)",
    "m82":  "Obstetric fistula (self-reported; 0=no, 1=yes)",
    # h-series: child health
    "h1":  "Has vaccination card",
    "h2":  "BCG vaccine received",
    "h3":  "BCG vaccine from card",
    "h4":  "DPT/Penta 1 received",
    "h5":  "DPT/Penta 1 from card",
    "h6":  "DPT/Penta 2 received",
    "h8":  "DPT/Penta 3 received",
    "h9":  "Measles vaccine received",
    "h10": "Polio 0 received",
    "h11": "Has any vaccination",
    "h12": "Polio 1 received",
    "h14": "Polio 2 received",
    "h16": "Polio 3 received",
    "h22": "Had fever in last 2 weeks",
    "h31": "Had cough in last 2 weeks",
    "h32": "Had diarrhoea in last 2 weeks",
    "h33": "Had diarrhoea in last 2 weeks (alt)",
    "h11": "Received any vaccination",
    # HW-specific
    "hw1": "Child's age in months",
    "hw2": "Child's weight (kg × 10)",
    "hw3": "Child's height (cm × 10)",
    "hw5": "Weight-for-age z-score (WAZ × 100; stunting proxy)",
    "hw8": "Height-for-age z-score (HAZ × 100; stunting)",
    "hw11":"Weight-for-height z-score (WHZ × 100; wasting)",
    "hw13":"BMI-for-age z-score",
    "hw70":"Height-for-age z-score (flagged)",
    "hw71":"Weight-for-age z-score (flagged)",
    "hw72":"Weight-for-height z-score (flagged)",
    # GE-specific
    "dhsclust":  "DHS cluster number (= v001 in IR/BR/HR)",
    "latnum":    "Cluster latitude (decimal degrees; 0 = displaced/missing)",
    "longnum":   "Cluster longitude (decimal degrees; 0 = displaced/missing)",
    "urban_rura":"Urban/rural classification of cluster (U=urban, R=rural)",
    "alt_dem":   "Altitude from digital elevation model (metres)",
    "dhsregna":  "Region name",
    "dhsregco":  "Region code",
    "dhscc":     "Country code (ISO2)",
    "dhsyear":   "Year of DHS survey",
    # AR-specific
    "hivclust":  "Cluster number (= v001; use to link AR to IR/MR)",
    "hivnumb":   "Household number (= v002)",
    "hivline":   "Line number (= v003)",
    "hiv03":     "HIV test result (0=negative, 1=positive, 7=indeterminate)",
    "hiv05":     "Sample weight for HIV sample",
    # WI-specific
    "whhid":     "Household ID for wealth index",
    "wlthindf":  "Wealth index factor score (continuous)",
    "wlthind5":  "Wealth quintile (1=poorest, 5=richest)",
    # NR p-series
    "p0":  "Pregnancy outcome code (0=live birth, see p1 for full classification)",
    "p1":  "Month of pregnancy outcome",
    "p2":  "Year of pregnancy outcome",
    "p3":  "Date of pregnancy outcome (CMC)",
    "p4":  "Completeness of date of pregnancy outcome (1=complete, 2=month imputed, 3=year imputed)",
    "p5":  "Child alive at time of survey (0=dead, 1=alive)",
    "p6":  "Age at death (CMC days, if died)",
    "p7":  "Source of birth weight info",
    "p8":  "Birth size",
    "p9":  "Gestational age at delivery",
    "p10": "Place of delivery",
    "p11": "Number of ANC visits",
    "p12": "Months pregnant at first ANC visit",
    "p13": "Blood pressure taken during ANC (0=no, 1=yes)",
    "p15": "Iron supplements received during pregnancy (0=no, 1=yes)",
    "p16": "Place of ANC",
    "p17": "Type of assistance at delivery",
    "p18": "Date of delivery",
    "p19": "Number of postnatal visits",
    "p20": "Who conducted postnatal check",
    "p21": "Date of postnatal check",
    "p30": "ANC from skilled/doctor provider (0=no, 1=yes)",
    "p31": "Complications during pregnancy (0=no, 1=yes)",
    "p32": "Type of delivery assistance",
    # BR-GEO enrichment columns
    "comp_any":                  "Any complication during delivery (0=no, 1=yes) [from GR m77]",
    "comp_hemorrhage":           "Excessive bleeding / haemorrhage (0=no, 1=yes) [from GR m78a]",
    "comp_eclampsia_convulsions":"Convulsions / eclampsia (0=no, 1=yes) [from GR m78b]",
    "comp_prolonged_labor":      "Prolonged or obstructed labour (0=no, 1=yes) [from GR m78c]",
    "comp_fever":                "Fever / infection during delivery (0=no, 1=yes) [from GR m78d]",
    "comp_fistula_symptoms":     "Obstetric fistula symptoms (0=no, 1=yes) [from GR m78e]",
    "comp_prom":                 "Premature rupture of membranes/PROM (0=no, 1=yes) [from GR m78f]",
    "comp_malaria":              "Malaria during pregnancy (0=no, 1=yes) [from GR m78g]",
    "comp_anemia":               "Anaemia (0=no, 1=yes) [from GR m78h]",
    "comp_hypertension":         "High blood pressure / hypertension (0=no, 1=yes) [from GR m78j]",
    "comp_other_a":              "Pre-eclampsia / gestational hypertension (0=no, 1=yes) [from GR m78m]",
    "comp_other_b":              "Antepartum haemorrhage (0=no, 1=yes) [from GR m78n]",
    "comp_other_c":              "Other specified complication (0=no, 1=yes) [from GR m78o]",
    "comp_postpartum":           "Complication after delivery (0=no, 1=yes) [from GR m55]",
    "comp_blood_transfusion":    "Received blood transfusion (0=no, 1=yes) [from GR m60]",
    "comp_pph":                  "Postpartum haemorrhage (0=no, 1=yes) [from GR m66]",
    "postnatal_care_timing":     "Timing of postnatal care (1=same day…5=42+ days) [from GR m80]",
    "fistula_reported":          "Obstetric fistula self-reported (0=no, 1=yes) [from GR m82]",
    "ir_had_terminated_pregnancy":"Mother had pregnancy not ending in live birth (1=yes) [from IR v228]",
    "ir_obstetric_fistula":      "Mother has obstetric fistula (1=yes) [from IR v234]",
    "ir_bp_taken_anc":           "Blood pressure taken during ANC for this pregnancy (0=no, 1=yes) [from IR v248]",
    "nr_outcome_code":           "Pregnancy outcome code [from NR p0]",
    "nr_child_alive_at_survey":  "Child alive at time of survey (0=dead, 1=alive) [from NR p5]",
    "nr_birth_size":             "Birth size [from NR p8]",
    "nr_gestational_age":        "Gestational age at delivery [from NR p9]",
    "nr_place_of_delivery":      "Place of delivery [from NR p10]",
    "nr_anc_visits":             "Number of ANC visits [from NR p11]",
    "nr_anc_months_pregnant":    "Months pregnant at 1st ANC [from NR p12]",
    "nr_anc_provider":           "ANC provider type [from NR p30]",
    "nr_complications":          "Complications during pregnancy (0=no, 1=yes) [from NR p31]",
    "nr_delivery_assistance":    "Type of delivery assistance [from NR p32]",
}

# Prefix lookup: v-series, hv-series, sh-series, mv-series, etc.
PREFIX_LABELS = {
    "hv001": "Cluster number (household survey)",
    "hv002": "Household number",
    "hv003": "Respondent line number",
    "hv005": "Household sample weight",
    "hv007": "Year of interview",
    "hv009": "Number of household members",
    "hv010": "Number of eligible women",
    "hv011": "Number of eligible men",
    "hv012": "Number of de jure members",
    "hv013": "Number of de facto members",
    "hv014": "Number of children under 5",
    "hv015": "Result of household interview",
    "hv016": "Day of interview",
    "hv021": "Primary sampling unit",
    "hv022": "Sample stratum",
    "hv023": "Stratification variable",
    "hv024": "Region",
    "hv025": "Urban/rural",
    "hv040": "Cluster altitude",
    "hv045": "Source of drinking water",
    "hv201": "Source of drinking water",
    "hv202": "Source of water for cooking",
    "hv204": "Time to get water",
    "hv205": "Type of toilet facility",
    "hv206": "Has electricity",
    "hv207": "Has radio",
    "hv208": "Has television",
    "hv209": "Has refrigerator",
    "hv210": "Has bicycle",
    "hv211": "Has motorcycle",
    "hv212": "Has car/truck",
    "hv213": "Main floor material",
    "hv214": "Main wall material",
    "hv215": "Main roof material",
    "hv216": "Number of rooms used for sleeping",
    "hv219": "Sex of household head",
    "hv220": "Age of household head",
    "hv221": "Has landline telephone",
    "hv225": "Has handwashing facility",
    "hv226": "Type of cooking fuel",
    "hv227": "Has mosquito bednet",
    "hv228": "Child under 5 slept under bednet",
    "hv230": "Water treatment method",
    "hv237": "Anything done to water to make it safe",
    "hv240": "Has watch/clock",
    "hv241": "Has computer",
    "hv242": "Has bank account",
    "hv243": "Has mobile phone",
    "hv244": "Owns agricultural land",
    "hv245": "Hectares of agricultural land",
    "hv246": "Owns livestock",
    "hv247": "Owns mobile phone (alt)",
    "hv270": "Wealth index quintile (1=poorest, 5=richest)",
    "hv271": "Wealth index factor score (continuous)",
}


def get_label(col):
    """Look up a human-readable label for a DHS column name."""
    # Direct match
    if col in LABELS:
        return LABELS[col]
    if col in PREFIX_LABELS:
        return PREFIX_LABELS[col]
    # Strip birth-index suffix (_1 to _6)
    base = re.sub(r'_[1-6]$', '', col)
    if base in LABELS:
        return LABELS[base] + f"  [for birth {col.split('_')[-1]}]"
    # hv-series
    if col in PREFIX_LABELS:
        return PREFIX_LABELS[col]
    # s-variables (country-specific)
    if col.startswith('s') and len(col) > 1 and col[1:].isdigit():
        return "Country-specific variable"
    if re.match(r'^s\d', col):
        return "Country-specific variable"
    # ml-series (malaria)
    if col.startswith('ml'):
        return "Malaria module variable"
    # ha-series (HIV / anthropometry)
    if col.startswith('ha'):
        return "HIV/anthropometry variable"
    # rc-series (reproductive calendar)
    if col.startswith('rc'):
        return "Reproductive calendar variable"
    return ""


def get_examples(con, table, col, dtype, limit=5):
    """Get a short list of example non-null values."""
    try:
        rows = con.execute(
            f'SELECT DISTINCT "{col}" FROM {table} WHERE "{col}" IS NOT NULL ORDER BY 1 LIMIT {limit}'
        ).fetchall()
        vals = [str(r[0]) for r in rows]
        return ", ".join(vals)
    except Exception:
        return ""


def pct_nonnull(con, table, col, total):
    try:
        n = con.execute(f'SELECT COUNT(*) FROM {table} WHERE "{col}" IS NOT NULL').fetchone()[0]
        return round(100.0 * n / total, 1) if total > 0 else 0
    except Exception:
        return 0


# Regex patterns for standard DHS core variables (fast name-based filter for wide tables)
CORE_PATTERNS = re.compile(r"""^(
    survey_code|country_code|country_name|dhs_phase|source_file|
    caseid|hhid|whhid|bidx\w*|bord\w*|pidx\w*|pord\w*|
    v\d{1,3}[a-z]?$|                         # v001-v999 (core women's)
    b\d{1,2}[a-z]?$|                         # b0-b25
    m\d{1,2}[a-z]?(_[1-6])?$|               # m1-m82 with optional birth suffix
    h\d{1,2}[a-z]?(_\d)?$|                  # h1-h99 child health
    hw\d{1,2}$|                              # hw1-hw99 anthropometry
    hv\d{1,3}[a-z]?$|                       # hv001-hv999 household
    sh\d{1,3}[a-z]?$|                       # sh (household country-specific, some standard)
    mv\d{1,3}[a-z]?$|                       # mv (men's)
    p\d{1,2}[a-z]?$|                        # p0-p32 pregnancy
    hivclust|hivnumb|hivline|hiv\d+[a-z]?$| # HIV
    wlthindf|wlthind5|dhsclust|latnum|longnum|urban_rura|alt_dem|
    dhsregna|dhsregco|dhscc|dhsyear|
    comp_\w+|nr_\w+|ir_\w+|postnatal\w+|fistula\w+
)$""", re.VERBOSE)


# ── Excel styling ─────────────────────────────────────────────────────────────
def make_header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

RECODE_COLORS = {
    "AR": "C9E6F0", "BR": "D5E8D4", "CR": "FFE6CC",
    "GE": "E1D5E7", "GR": "FFD7BE", "HR": "DAE8FC",
    "HW": "FFF2CC", "IR": "D5E8D4", "KR": "F8CECC",
    "MR": "D4E1F7", "NR": "E6D0DE", "PR": "DAE8FC",
    "SR": "F0F0F0", "WI": "FFF9C4", "BR_GEO": "B7E1CD",
}

def style_header_row(ws, row_num, fill_hex):
    fill = make_header_fill(fill_hex)
    bold = Font(bold=True)
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = fill
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("00_Summary")
    ws_sum.append(["Code", "Recode Name", "Description", "File (MB)", "Rows", "Total Cols", "Documented Cols"])
    style_header_row(ws_sum, 1, "4A86C8")
    for cell in ws_sum[1]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF")

    for code, (dbfile, table, name, desc) in RECODES.items():
        dbpath = f"{BASE}/{dbfile}"
        if not os.path.exists(dbpath):
            ws_sum.append([code, name, desc, "—", "—", "—", "—"])
            continue
        con = duckdb.connect(dbpath, read_only=True)
        total_rows = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        cols = [r[0] for r in con.execute(f"DESCRIBE {table}").fetchall()]
        size_mb = round(os.path.getsize(dbpath) / 1e6, 1)
        doc = sum(1 for c in cols if get_label(c))
        ws_sum.append([code, name, desc, size_mb, total_rows, len(cols), doc])
        con.close()

    for col_idx, width in enumerate([8, 35, 60, 10, 15, 12, 14], start=1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Per-recode sheets ───────────────────────────────────────────────────
    for code, (dbfile, table, name, desc) in RECODES.items():
        dbpath = f"{BASE}/{dbfile}"
        if not os.path.exists(dbpath):
            print(f"  SKIP {code}: file not found")
            continue

        print(f"  Processing {code} ({name}) ...", flush=True)
        con = duckdb.connect(dbpath, read_only=True)
        total_rows = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        describe = con.execute(f"DESCRIBE {table}").fetchall()
        all_cols = [(r[0], r[1]) for r in describe]

        # For wide databases: filter by standard DHS variable name patterns
        # (avoids running thousands of COUNT queries — name filter is instant)
        if len(all_cols) > 500:
            print(f"    Wide table ({len(all_cols)} cols) — filtering by core variable names ...", flush=True)
            core = [(col, dtype) for col, dtype in all_cols if CORE_PATTERNS.match(col)]
            print(f"    → {len(core)} core cols kept")
            # Sample non-null % for kept columns only (much smaller set)
            keep = []
            for col, dtype in core:
                pct = pct_nonnull(con, table, col, total_rows)
                keep.append((col, dtype, pct))
        else:
            keep = []
            for col, dtype in all_cols:
                pct = pct_nonnull(con, table, col, total_rows)
                keep.append((col, dtype, pct))

        ws = wb.create_sheet(code)
        # Sheet title row
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{code}  —  {name}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].fill = make_header_fill(RECODE_COLORS.get(code, "EEEEEE"))
        ws.merge_cells("A2:F2")
        ws["A2"] = desc
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 30

        ws.append([])  # blank row 3
        ws.append(["Variable", "DHS Label / Description", "Data Type", "% Non-Null", "Example Values"])
        style_header_row(ws, 4, RECODE_COLORS.get(code, "CCCCCC"))

        # Alternating row colors
        light = PatternFill("solid", fgColor="F7F7F7")
        for i, (col, dtype, pct) in enumerate(keep):
            label = get_label(col)
            ex = get_examples(con, table, col, dtype) if pct > 0 else ""
            row = [col, label, dtype, f"{pct}%", ex]
            ws.append(row)
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = light
            # Bold label for well-known variables
            if label and not label.startswith("Country"):
                ws.cell(ws.max_row, 1).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 30
        ws.freeze_panes = "A5"

        # Footer: total / filtered
        ws.append([])
        ws.append([f"Total columns in database: {len(all_cols)}   |   Shown here: {len(keep)}   |   Total rows: {total_rows:,}"])

        con.close()
        print(f"    → {len(keep)} rows written to sheet {code}")

    wb.save(OUT_EXCEL)
    print(f"\nSaved: {OUT_EXCEL}")


if __name__ == "__main__":
    main()
